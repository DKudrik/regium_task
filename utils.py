from openpyxl import Workbook

from db_utils import Advertisement


def create_xls(session) -> Workbook:
    """ Creates xls file. Then in a for loop creates a header and in the
    second loop adss data from the DB"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cars"
    columns = [
        "advt_id",
        "price",
        "price_eur",
        "year",
        "mileage",
        "engine_volume",
        "transmission",
        "horse_power",
        "drive_wheels",
        "fuel",
        "is_market_price",
        "is_only_on_avito",
        "is_owner",
        "is_damaged",
        "description",
        "place_of_sale",
        "url_to_advt_page",
        "created",
    ]
    for idx, column in enumerate(columns, 1):
        ws.cell(row=1, column=idx, value=column)
    advts = session.query(Advertisement).all()
    for idx, advt in enumerate(advts, 2):
        for col_idx, col in enumerate(columns, 1):
            ws.cell(row=idx, column=col_idx, value=vars(advt)[col])
    wb.save("avito_cars.xlsx")
